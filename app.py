import streamlit as st
from st_social_media_links import SocialMediaIcons
from streamlit.components.v1 import html as st_html
from streamlit_option_menu import option_menu
from streamlit_modal import Modal

import json
import pandas as pd
import numpy as np
import calendar
from datetime import datetime
import os
from io import BytesIO

import locale
import base64
import zipfile

import openpyxl
from spire.xls import *
from spire.xls.common import *

import git
import subprocess    


#Questo è il file su cui lavorare
st.set_page_config(
    page_title="ASA - Stagione 2024/25",
    page_icon="logo_2.ico",  
    layout='wide',  
)

# ---------- MAIN PAGE ----------
col1, col2 = st.columns([5, 20])
with col1:
    st.image('logo.png', width=60)
with col2:
    st.title('Athletic Soccer Academy')

st.html("""
        <hr>
        """)

def load_credentials(file_path):
    with open(file_path, 'r') as file:
        return json.load(file)
    


#locale.setlocale(locale.LC_TIME, 'it_IT.UTF-8')
def converti_data(data_string):
    # Converte la stringa in un oggetto datetime
    dt = datetime.strptime(data_string, "\"%Y-%m-%dT%H:%M:%S\"")
    # Formatta la data nel formato desiderato
    data_formattata = dt.strftime("%A %d-%m-%Y ore %H:%M")
    return data_formattata


def download_link(object_to_download, download_filename, download_link_text):
    """
    Genera un link per il download di un oggetto binario.
    """
    if isinstance(object_to_download, bytes):
        b64 = base64.b64encode(object_to_download).decode()
    else:
        object_to_download = object_to_download.encode()

    b64 = base64.b64encode(object_to_download).decode()
    return f'<a href="data:file/txt;base64,{b64}" download="{download_filename}">{download_link_text}</a>'

# Funzione per caricare le credenziali dal file JSON
def carica_credenziali(file_path):
    with open(file_path, 'r') as file:
        return json.load(file)

# Caricamento delle credenziali
credenziali = carica_credenziali('Mister.json')

#st.write(credenziali)
# Funzione per verificare le credenziali
def verifica_credenziali(username, password):
    for allenatore in credenziali["allenatore"]:
        if allenatore["mister"] == username and allenatore["pwd"] == password:
            return True
    return False


def convert_df_to_excel(df, file_path):
    df.to_excel(file_path, index=False)

def serialize_datetime(obj): 
    if isinstance(obj, datetime): 
        return obj.isoformat() 
    raise TypeError("Type not serializable") 

def mostra_contenuto_json(file_path):
    with open(file_path, 'r') as file:
        data = json.load(file)
    st.json(data)

# Funzione per visualizzare e modificare il contenuto dei file JSON in una finestra modale
def mostra_e_modifica_json(file_path, directory, cat):
    file_complete = os.path.basename(file_path)
    nfile, _ = os.path.splitext(file_complete)

    with open(file_path, 'r') as file:
        data = json.load(file)
    
    st.subheader("Modifica Convocazione")
    
    with st.expander(f"Modifica {file_path}", expanded=True):
        # Aggiungi input per modificare i campi esistenti
        data['squadra'] = st.text_input("Squadra", data['squadra'])
        data['data_ora_incontro'] = st.text_input("Data e Ora Incontro", data['data_ora_incontro'])
        data['denominazione_campo'] = st.text_input("Denominazione Campo", data['denominazione_campo'])
        data['ora_raduno'] = st.text_input("Ora Raduno", data['ora_raduno'])

        non_convocati_list = data['non_convocati'].split(', ')

        # Rimuovi i componenti attuali dalla lista non_convocati
        non_convocati_list = [giocatore for giocatore in non_convocati_list if giocatore not in data['componenti_squadra']]

        componenti_squadra_modificati = []
        for i in range(0, len(data['componenti_squadra'])):
            giocatore_attuale = data['componenti_squadra'][i]
            options = [giocatore_attuale] + non_convocati_list
            index = options.index(giocatore_attuale) if giocatore_attuale else 0
            giocatore_selezionato = st.selectbox(f"Giocatore {i+1}", options=options, index=index, key=f'componente_{i+1}')
            
            if giocatore_selezionato != giocatore_attuale:
                if giocatore_attuale:
                    non_convocati_list.append(giocatore_attuale)
                if giocatore_selezionato:
                    non_convocati_list.remove(giocatore_selezionato)
            
            componenti_squadra_modificati.append(giocatore_selezionato)

        data['componenti_squadra'] = componenti_squadra_modificati
        data['non_convocati'] = ', '.join(non_convocati_list)
        
        st.text_area("Non Convocati", value=data['non_convocati'], key='non_convocati_1')

        data['nome_mister'] = st.text_input("Nome Mister", data['nome_mister'])
        data['nome_dirigente'] = st.text_input("Nome Dirigente", data['nome_dirigente'])
        
        if st.button("Salva Modifiche"):
            with open(file_path, 'w') as file:
                json.dump(data, file, indent=4)
            st.success("Convocazione modificata con successo!")
    
            # Aggiungi la logica per il pulsante "Salva convocazione"
            with open(file_path, 'r') as file:
                data = json.load(file)
            
            convocazione = {
                    "giornata": data['giornata'],
                    "squadra": data['squadra'],
                    "data_ora_incontro": data['data_ora_incontro'],
                    "denominazione_campo": data['denominazione_campo'],
                    "ora_raduno": data['ora_raduno'],
                    "componenti_squadra": data['componenti_squadra'],
                    "non_convocati": data['non_convocati'],
                    "nome_mister": data['nome_mister'],
                    "nome_dirigente": data['nome_dirigente']
            }
            # Esporta la convocazione in Excel
            name_xlsx = directory + nfile + '.xlsx'
            try:
                workbook = openpyxl.load_workbook(name_xlsx)
            except FileNotFoundError:
                workbook = openpyxl.Workbook()


            txt_cell = f"CONVOCAZIONI {cat}"
            sheet = workbook.active
            sheet['C8'] = txt_cell
            sheet['C10'] = data['squadra']
            match_day = converti_data(data['data_ora_incontro'])
            sheet['C14'] = match_day
            parti = data['denominazione_campo'].split(',')
            for p, parte in enumerate(parti):
                cell = f'C{16 + p}'
                sheet[cell] = parte.strip()
            sheet['C19'] = data['ora_raduno']
            # Inserisci i nomi dei componenti della squadra nelle celle C22:C43

            for i, giocatore in enumerate(data['componenti_squadra']):
                cell = f'C{22 + i}'
                sheet[cell] = giocatore
            sheet['C45'] = data['non_convocati']
            sheet['C50'] = data['nome_mister']
            sheet['C52'] = data['nome_dirigente']

            # Salva il file in un oggetto BytesIO
            file_name = nfile + '.xlsx'
            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            # Salva il file Excel nella cartella convocazioni
            #fname = directory + nome_file + '.xlsx'
            with open(name_xlsx, 'wb') as f:
                f.write(buffer.getvalue())

            # # Converti il file Excel in PDF
            # pdf_file_name = f'{name_xlsx}.pdf'
            # workbook_spire = Workbook()
            # workbook_spire.LoadFromFile(name_xlsx)
            # workbook_spire.ConverterSetting.SheetFitToPage = True
            # workbook_spire.SaveToFile(pdf_file_name, FileFormat.PDF)
            # workbook_spire.Dispose()
            commit_message = "Update file"
            # commit_and_push(repo, commit_message, token, credentials_path)
            # Genera il link di download per il excel
            with open(name_xlsx, 'rb') as f:
                excel_buffer = BytesIO(f.read())
            download_link_html = download_link(excel_buffer.getvalue(), name_xlsx, 'Clicca qui per scaricare il file Excel')
            st.markdown(download_link_html, unsafe_allow_html=True)



# Definizione delle pagine
def homepage():
    pass

def aggiungi_giocatore():
    nuovo_numero = df['N.'].max() + 1  # Incrementa il valore di N.
    nuovo_nome = st.session_state['nuovo_nome']
    nuovo_cognome = st.session_state['nuovo_cognome']
    nuovo_anno = st.session_state['nuovo_anno']
    nuovo_ruolo = st.session_state['nuovo_ruolo']
    
    # Crea un nuovo DataFrame con il nuovo giocatore
    nuovo_giocatore = pd.DataFrame({
        'N.': [nuovo_numero],
        'NOME': [nuovo_nome],
        'COGNOME': [nuovo_cognome],
        'ANNO': [nuovo_anno],
        'RUOLO': [nuovo_ruolo]
    })
    
    # Aggiungi il nuovo giocatore al DataFrame esistente
    df = pd.concat([df, nuovo_giocatore], ignore_index=True)
    
    # Salva il DataFrame aggiornato nel file CSV
    df.to_csv('U16P.csv', index=False, sep=';')
    
    st.success("Giocatore aggiunto con successo!")
    st.table(df)

# _DIR = "/wokspaces/ASA_SGS/"
# _DIR_CS = "/workspaces/ASA_SGS/data"

def get_mister_info(username, mister_data):
    for allenatore in mister_data['allenatore']:
        if allenatore['mister'] == username:
            return allenatore
    return None

repo = 0

def gestione_rosa():
    # Inserimento credenziali per la gestione
    st.title("Gestione della Rosa")
    # Accesso a github


    if repo==0:
        # Form di login
        if "logged_in" not in st.session_state:
            st.session_state.logged_in = False
        if not st.session_state.logged_in:
            st.subheader("Login")
            username = st.selectbox("Seleziona il tuo Gruppo Squadra", 
                                ("Prima Squadra", "Under 19", "Under 18", "Under 17R", "Under 17P", "Under 16R", "Under 16P", "Under 15R", "Under 15P", "Under 14R", "Under 14P"),
                                    index = None,
                                    placeholder="seleziona... ")
            password = st.text_input("Password", type="password")
            if st.button("Accedi"):
                if verifica_credenziali(username, password):
                    st.session_state.logged_in = True
                    st.session_state.username = username
                    st.rerun()
                else:
                    st.error("Credenziali errate. Riprova.")

        else:
            
            mister_info = get_mister_info(st.session_state.username, credenziali)
            #st.write(f"Accesso Effettuato, {st.session_state.username}")
            if mister_info:
                    st.write(f"Accesso Effettuato, {st.session_state.username}")
                    # assegnare il nome del file
                    if st.button("Logout"):
                                st.session_state.logged_in = False
                                st.rerun()

                    df = pd.read_csv(mister_info['file'], delimiter=';')  # Specifica il delimitatore
                    registro = mister_info['registro']
                    minuti = mister_info['minuti']
                    acronimo = 'Convocazioni/' + mister_info['acronimo'] + '/'
                    reportistica = 'Campionato/' + mister_info['acronimo'] + '/'
            else:
                    st.write("Informazioni del mister non trovate.")

            selected = option_menu(None, ["Lista", "Presenze",  "Archivio Presenze", "Convocazione", "Archivio Convocazioni", "Report Partita", "Reportistica"], 
                icons=['list', 'calendar-check', 'calendar-check', 'clipboard-check', 'clipboard-check', 'file-text', 'file-text'],            
                menu_icon="cast", 
                default_index=0, 
                orientation="horizontal",
                styles={
                    "container": {"padding": "0!important", "background-color": "#000000"},
                    "icon": {"color": "orange", "font-size": "14px"}, 
                    "nav-link": {"font-size": "14px", "text-align": "left", "margin":"0px", "--hover-color": "#eee"},
                    "nav-link-selected": {"background-color": "green"},
                }
            )

            if repo==0:
                if selected == "Lista":


                    
                    # Apertura del file CSV e visualizzazione della tabella ordinata per cognome
                    try:
                        #df = pd.read_csv(_DIR + 'U16P.csv', delimiter=';')  # Specifica il delimitatore
                        edited_df = st.data_editor(df, num_rows="dynamic", use_container_width=True)
                        #st.table(df)
                    except Exception as e:
                        st.error(f"Errore durante la lettura del file CSV: {e}")

                    @st.dialog("Aggiungi giocatore")

                    def add_player(df):

                        st.text_input("Nome", key='nuovo_nome')
                        st.text_input("Cognome", key='nuovo_cognome')
                        st.number_input("Anno di nascita", min_value=1900, max_value=datetime.now().year, step=1, key='nuovo_anno')
                        st.selectbox("Ruolo", options=["P", "D", "C", "A"], key='nuovo_ruolo', placeholder="Ruolo")


                        # st.write(f"Why is {item} your favorite?")
                        # reason = st.text_input("Because...")
                        if st.button("Conferma"):
                            nuovo_nome = st.session_state['nuovo_nome']
                            nuovo_cognome = st.session_state['nuovo_cognome']
                            nuovo_anno = st.session_state['nuovo_anno']
                            nuovo_ruolo = st.session_state['nuovo_ruolo']
                            # Crea un nuovo DataFrame con il nuovo giocatore
                            nuovo_giocatore = pd.DataFrame({
                                'NOME': [nuovo_nome],
                                'COGNOME': [nuovo_cognome],
                                'ANNO': [nuovo_anno],
                                'RUOLO': [nuovo_ruolo]
                            })
                            
                            # Aggiungi il nuovo giocatore al DataFrame esistente
                            df = pd.concat([df, nuovo_giocatore], ignore_index=True)
                            
                            # Salva il DataFrame aggiornato nel file CSV
                            df.to_csv(mister_info['file'], index=False, sep=';')

                            st.success("Giocatore aggiunto con successo!")
                            #st.rerun()



                    # if "add_player" not in st.session_state:
                    #     if st.button("Aggiungi giocatore"):
                    #         add_player(df)
                    # else:
                    #     f"Non è stato possibile aggiungere {st.session_state.add_player['item']}, in quanto: {st.session_state.add_player['reason']}"

                    # Add a button to save changes

                    st.write(":blue[Aggiungi/Modifica/Elimina un giocatore]")
                    row_1 = '''
                        Puoi modificare un giocatore direttamente nella tabella e cliccando, successivamente, sul pulsante 'Salva modifiche'.
                    '''
                    row_2 = '''
                        Puoi eliminare un giocatore andando a spuntarlo, in corrispondenza del quadratino alla sua sinistra, e cliccando sull'icona 'cestino' che 
                        compare in alto a destra. Per salvare le modifiche cliccare su 'Salva modifiche'. 
                    '''
                    row_3 = '''
                        Puoi aggiungere un nuovo giocatore direttamente nella tabella e cliccando, successivamente, sul pulsante 'Salva modifiche'.
                    '''
                    st.markdown(row_3)
                    st.markdown(row_1)
                    st.markdown(row_2)

                    if st.button('Salva modifiche'):
                        edited_df.to_csv(mister_info['file'], sep=";", index=False)
                        st.success('Modifica effettuata!')
                        #st.session_state.df = edited_df

                        commit_message = "Update file"
                        # Impostazioni della funzione
                        # Esegui la funzione
                        # commit_and_push(repo, commit_message, token, credentials_path)
                        # commit_and_push(repo, commit_message)       
                        #st.rerun()
                    

                elif selected == "Presenze":
                    # Visualizzazione del calendario e gestione presenze
                    st.subheader("Calendario Presenze")
                    giorno = st.date_input("", datetime.today(), format="DD/MM/YYYY")

                    try:
                        #df = pd.read_csv(_DIR + 'U16P.csv', delimiter=';')  # Specifica il delimitatore
                        presenze = []
                        
                        # Creazione della tabella per l'inserimento delle presenze
                        col1, col2, col3, col4 = st.columns([1, 2, 2, 2])
                        col1.text("ID")
                        col2.write("NOME")
                        col3.write("COGNOME")
                        col4.write("TIPO")

                        
                        for index, row in df.iterrows():
                            col1, col2, col3, col4 = st.columns([1, 2, 2, 2])
                            col1.write(index)
                            col2.write(row['NOME'])
                            col3.write(row['COGNOME'])
                            presenza = col4.selectbox("Presenza", ("P", "A", "INF", "MAL", "IND"), key=index, label_visibility='collapsed')
                            presenze.append({
                                "id": index,
                                "nome": row['NOME'],
                                "cognome": row['COGNOME'],
                                "presente": presenza
                            })
                        
                        if st.button("Salva Presenze", icon="💾"):
                            filename = registro
                            new_data = {
                                giorno.strftime("%Y-%m-%d"): {
                                    "presenze": presenze
                                }
                            }
                            
                            if os.path.exists(filename):
                                with open(filename, 'r+') as file:
                                    file_data = json.load(file)
                                    # Sovrascrivi i dati del giorno esistente o aggiungi nuovi dati
                                    file_data.update(new_data)
                                    file.seek(0)
                                    json.dump(file_data, file, indent=4)
                            else:
                                with open(filename, 'w') as file:
                                    json.dump(new_data, file, indent=4)

                            st.success("Presenze salvate con successo!")
                    except Exception as e:
                        st.error(f"Errore durante la lettura del file CSV: {e}") 

                elif selected=="Archivio Presenze":
                    try:
                        with open(registro, 'r') as file:
                            presenze_data = json.load(file)
                        
                        # Creazione del menu di navigazione per i mesi e gli anni
                        anni_mesi_disponibili = sorted(set(data_giorno[:7] for data_giorno in presenze_data.keys()))
                        anni_mesi_disponibili_nomi = [f"{calendar.month_name[int(m.split('-')[1])]} {m.split('-')[0]}" for m in anni_mesi_disponibili]
                        anno_mese_corrente = datetime.today().strftime("%Y-%m")
                        anno_mese_corrente_nome = f"{calendar.month_name[int(anno_mese_corrente.split('-')[1])]} {anno_mese_corrente.split('-')[0]}"
                        anno_mese_selezionato = st.selectbox("Seleziona il mese e l'anno da visualizzare", anni_mesi_disponibili_nomi, index=anni_mesi_disponibili_nomi.index(anno_mese_corrente_nome))     

                        # Visualizzazione dei dati per il mese e l'anno selezionato
                        st.subheader(f"Presenze del mese di {anno_mese_selezionato}")
                        anno_mese_selezionato_numero = anni_mesi_disponibili[anni_mesi_disponibili_nomi.index(anno_mese_selezionato)]   

                        # Popolamento dei dati nella tabella
                        giocatori = {giocatore['id']: giocatore for data_giorno, dettagli in presenze_data.items() for giocatore in dettagli["presenze"]}
                        data = []
                        
                        for data_giorno, dettagli in presenze_data.items():
                            if data_giorno.startswith(anno_mese_selezionato_numero):
                                giorno = int(data_giorno.split("-")[2])  # Estrai il giorno dalla data
                                for presenza in dettagli["presenze"]:
                                    if presenza['id'] in giocatori:
                                        giocatori[presenza['id']][f"{giorno}"] = presenza['presente']
                        
                        for giocatore in giocatori.values():
                            row = [giocatore['nome'], giocatore['cognome']] + [giocatore.get(f"{i}", "") for i in range(1, 32)]
                            data.append(row)
                        
                        columns = ["Nome", "Cognome"] + [f"{i}" for i in range(1, 32)]
                        df_presenze = pd.DataFrame(data, columns=columns)
                        st.table(df_presenze)
                        
                        # Controlla se il file Excel esiste già
                        # excel_file_path = 'presenze_giocatori.xlsx'
                        # if not os.path.exists(excel_file_path):
                        excel_file_path = f"presenze_giocatori_{anno_mese_selezionato_numero}.xlsx"
                        #     st.text(excel_file_path)
                        convert_df_to_excel(df_presenze, excel_file_path)

                        # Aggiungi il pulsante per il download in formato Excel
                        with open(excel_file_path, 'rb') as file:
                            st.download_button(
                                label="Scarica in formato Excel",
                                data=file,
                                file_name=excel_file_path,
                                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                            )
                    except Exception as e:
                        st.error(f"Errore durante la lettura del file JSON: {e}")

                elif selected=="Convocazione":
                    try:
                        # Leggi il file CSV esistente
                        #df = pd.read_csv(_DIR + 'U16P.csv', delimiter=';')  # Specifica il delimitatore
                        giocatori = [f"{row['NOME']} {row['COGNOME']}" for index, row in df.iterrows()]

                        # Creazione del form
                        st.subheader("Crea una nuova convocazione")
                        html = """
                                <div class="warning" style='padding:0.1em; background-color:#E9D8FD; color:#69337A'>
                                <span>
                                <p style='margin-top:1em; text-align:left'>
                                <b>NOTA: </b>
                                Se vuoi convocare un giocatore che non fa parte della tua rosa, vai nella sezione <b>Lista</b> e clicca sul pulsante <b>Aggiungi Giocatore</b>.
                                </p>
                                </span>
                                </div>
                                """
                        st.html(html)                
                        giornata = st.number_input("Giornata numero", value=0)
                        squadra = st.text_input("Squadra Avversaria")
                        #data_ora_incontro = st.text_input("Data e ora incontro", value=datetime.now().strftime("%Y-%m-%d %H:%M"))
                        # Input per data e ora dell'incontro
                        data_incontro = st.date_input("Data incontro", value=datetime.now().date())
                        ora_incontro = st.time_input("Ora incontro")
                        data_ora_incontro = datetime.combine(data_incontro, ora_incontro)
                        data_json = json.dumps(data_ora_incontro, default=serialize_datetime)
                        
                        # Formattazione della data e dell'ora
                        data_formattata = data_ora_incontro.strftime("%d/%m/%y")
                        ora_formattata = data_ora_incontro.strftime("%H:%M")


                        
                        denominazione_campo = st.text_input("Denominazione e indirizzo campo (formato 'Denominazione campo, indirizzo completo')")
                        ora_raduno = st.time_input("Ora raduno")


                        
                        # Selezione dei componenti della squadra
                        componenti_squadra = []
                        giocatori_disponibili = giocatori.copy()
                    
                        for i in range(20):
                            giocatore = st.selectbox(f"Giocatore {i+1}", options=[""] + giocatori_disponibili, index=0, key=f'componente_{i+1}')
                            if giocatore:
                                componenti_squadra.append(giocatore)
                                giocatori_disponibili.remove(giocatore)
                            else:
                                componenti_squadra.append("")

                        # for i in range(20):
                        #     giocatore = st.selectbox(f"Giocatore {i+1}", options=[""] + giocatori_disponibili, index=0, key=f'componente_{i+1}')
                        #     if giocatore:
                        #         componenti_squadra.append(giocatore)
                        #         giocatori_disponibili.remove(giocatore)
                        #non_convocati = st.text_area("Non convocati", value=", ".join(giocatori_disponibili), key='non_convocati_1')


                        # Aggiorna la lista dei non convocati
                        non_convocati = ", ".join(giocatori_disponibili)
                        st.text_area("Non convocati", value=non_convocati, key='non_convocati_1')
                        
                        nome_mister = st.text_input("Nome del mister")
                        nome_dirigente = st.text_input("Nome del dirigente")
                        
                        nome_file = st.text_input("Inserisci il nome del file (senza estensione):")
                        nome_file = f'{giornata}_{nome_file}'
                        

                        if st.button("Salva convocazione"):
                            convocazione = {
                                "giornata": giornata,
                                "squadra": squadra,
                                "data_ora_incontro": data_json,
                                "denominazione_campo": denominazione_campo,
                                "ora_raduno": str(ora_raduno),
                                "componenti_squadra": componenti_squadra,
                                "non_convocati": non_convocati,
                                "nome_mister": nome_mister,
                                "nome_dirigente": nome_dirigente
                            }
                            if nome_file:
                                nomefile = acronimo + nome_file
                                with open(f'{nomefile}.json', 'w') as file:
                                    json.dump(convocazione, file, indent=4)
                                
                                #Esporta la convocazione in excel
                                #Carica il file Excel esistente o creane uno nuovo
                                try:
                                    workbook = openpyxl.load_workbook('Convocazioni.xlsx')
                                except FileNotFoundError:
                                    workbook = openpyxl.Workbook()
                                # Rimuovi eventuali fogli di lavoro vuoti
                                for sheet in workbook.sheetnames:
                                    if workbook[sheet].max_row == 1 and workbook[sheet].max_column == 1:
                                        del workbook[sheet]
                                txt_cell = f"CONVOCAZIONI {mister_info['mister']}"
                                
                                sheet = workbook.active
                                sheet['C8'] = txt_cell
                                sheet['C10'] = squadra

                                match_day = converti_data(data_json)
                                sheet['C14'] = match_day

                                parti = denominazione_campo.split(',')

                                for p, parte in enumerate(parti):
                                    cell = f'C{16 + p}'
                                    sheet[cell] = parte.strip()

                                sheet['C19'] = ora_raduno
                                # Inserisci i nomi dei componenti della squadra nelle celle D19:D40
                                for i, giocatore in enumerate(componenti_squadra):
                                    cell = f'C{22 + i}'
                                    sheet[cell] = giocatore

                                sheet['C45'] = non_convocati
                                sheet['C50'] = nome_mister
                                sheet['C52'] = nome_dirigente

                                
                                # Salva il file in un oggetto BytesIO
                                file_name = nome_file + '.xlsx'
                                st.write(file_name)
                                buffer = BytesIO()
                                workbook.save(buffer)
                                buffer.seek(0)


                                
                                # Salva il file Excel nella cartella convocazioni
                                fname = acronimo + nome_file + '.xlsx'
                                with open(fname, 'wb') as f:
                                    f.write(buffer.getvalue())


                                # Converti il file Excel in PDF

                                # pdf_file_name = f'{fname}.pdf'
                                # workbook_spire = Workbook()
                                # workbook_spire.LoadFromFile(fname)
                                # workbook_spire.ConverterSetting.SheetFitToPage = True
                                # workbook_spire.SaveToFile(pdf_file_name, FileFormat.PDF)
                                # workbook_spire.Dispose()

                                # Genera il link di download per il file Excel
                                with open(fname, 'rb') as f:
                                    excel_buffer = BytesIO(f.read())
                                download_link_html = download_link(excel_buffer.getvalue(), fname, 'Clicca qui per scaricare il file Excel')
                                commit_message = "Update file"
                                st.success("Convocazione salvata con successo!")

                                st.markdown(download_link_html, unsafe_allow_html=True)
                                
                                # # Genera il link di download per il PDF
                                # with open(pdf_file_name, 'rb') as f:
                                #     pdf_buffer = BytesIO(f.read())
                                # download_link_html = download_link(pdf_buffer.getvalue(), pdf_file_name, 'Clicca qui per scaricare il file PDF')
                                # st.markdown(download_link_html, unsafe_allow_html=True)
                                

                            else:
                                st.error("Per favore, inserisci un nome per il file.")

                        # if st.button("Esporta in PDF"):
                        #     if nome_file:
                        #         pdf = PDF()
                        #         pdf.add_page()
                        #         pdf.chapter_title("Convocazione Gara")
                        #         pdf.chapter_body(f"Squadra Avversaria: {squadra}\nData Incontro: {data_formattata}\nOra Incontro: {ora_formattata}\nDenominazione Campo: {denominazione_campo}\nOra Raduno: {str(ora_raduno)}\nComponenti Squadra: {', '.join(componenti_squadra)}\nNon Convocati: {non_convocati}\nNome Mister: {nome_mister}\nNome Dirigente: {nome_dirigente}")
                        #         pdf.output(f"{nome_file}.pdf")
                        #         st.success("Convocazione esportata in PDF con successo!")
                        #     else:
                        #         st.error("Per favore, inserisci un nome per il file.")


                    except Exception as e:
                        st.error(f"Errore durante la lettura del file CSV: {e}")

                elif selected=="Archivio Convocazioni":

                        st.subheader("Archivio Convocazioni")
                        #directory = 'Convocazioni/U16P/'

                        # Elenca tutti i file nella directory
                        files = [f for f in os.listdir(acronimo) if f.endswith('.json')]
                                    
                        if len(files) > 0:
                            # Crea una selectbox per selezionare il file
                            files.insert(0, "")
                            selected_file = st.selectbox("Seleziona un file", files)
                            if selected_file:
                                # Crea il percorso completo del file selezionato
                                file_path = os.path.join(acronimo, selected_file)
                                    
                                # Mostra e modifica il contenuto del file JSON selezionato
                                st.write(f"Visualizza e modifica il contenuto del file: {file_path}")
                                mostra_e_modifica_json(file_path, acronimo, mister_info['mister'])
                                with open(file_path, "rb") as file:
                                    btn = st.download_button(
                                        label="Scarica la convocazione",
                                        data=file,
                                        file_name=selected_file,
                                        mime="application/json"
                                    )                        
                        else:
                            st.subheader("Non ci sono convocazioni")
                            
                            # Crea le colonne per i bottoni affiancati
                        #     cols = st.columns(len(files))
                        #     for idx, file in enumerate(files):
                        #         file_path = os.path.join(acronimo, file)
                        #         # Crea un link cliccabile per ciascun file in una colonna
                        #         if cols[idx].button(file):
                        #             st.session_state.selected_file = file_path
                            
                        #     # Mostra e modifica il contenuto del file JSON selezionato
                        #     if 'selected_file' in st.session_state:
                        #         st.write(f"Visualizza e modifica il contenuto del file: {st.session_state.selected_file}")
                        #         mostra_e_modifica_json(st.session_state.selected_file, acronimo)
                            
                        #     # # Mostra il contenuto del file JSON selezionato in una finestra modale
                        #     # if 'selected_file' in st.session_state:
                        #     #     mostra_contenuto_json(st.session_state.selected_file)
                        # else:
                        #     st.subheader("Non ci sono convocazioni")

                elif selected=="Report Partita":

                    def report_json(file_path, directory):
                        with open(file_path, 'r') as f:
                            data = json.load(f)
                        
                        st.json(data)

                    st.subheader("Report Partita")
                
                    # Elenca tutti i file nella directory
                    files = [f for f in os.listdir(reportistica) if f.endswith('.json')]
                        
                    # Crea le colonne per i bottoni affiancati
                    #cols = st.columns(len(files))

                    if len(files) > 0: 
                        # Crea una selectbox per selezionare il file
                        files.insert(0, "")
                        selected_file = st.selectbox("Visualizza un report già presente", files, key='report')
                        if selected_file:
                            # Crea il percorso completo del file selezionato
                            file_path = os.path.join(reportistica, selected_file)
                                
                            # Mostra e modifica il contenuto del file JSON selezionato
                            st.write(f"Visualizza e modifica il contenuto del file: {file_path}")
                            report_json(file_path, reportistica)
                            # Aggiungi un pulsante per il download del file
                            with open(file_path, "rb") as file:
                                btn = st.download_button(
                                    label="Scarica il report",
                                    data=file,
                                    file_name=selected_file,
                                    mime="application/json"
                                )

                        # cols = st.columns(len(files))
                        # for idx, file in enumerate(files):
                        #     file_path = os.path.join(acronimo, file)
                        #     # Crea un link cliccabile per ciascun file in una colonna
                        #     if cols[idx].button(file):
                        #             st.session_state.selected_file = file_path

                        # if 'selected_file' in st.session_state:
                        #     st.write(f"Visualizza e modifica il contenuto del file: {st.session_state.selected_file}")

                        #     st.write(st.session_state.selected_file)
                        #     with open(st.session_state.selected_file, 'r') as file:
                        #         convocazione = json.load(file)
                    st.subheader("Crea nuovo Report")
                    if st.expander("Nuovo Report"):
                        files_convocazione = [f for f in os.listdir(acronimo) if f.endswith('.json')]
                        # Trova i file non comuni ad entrambe le liste
                        non_comuni = list(set(files).symmetric_difference(set(files_convocazione)))
                        files_convocazione.insert(0, "")
                        selected_convocazione = st.selectbox("Seleziona la convocazione", non_comuni, key='convocazioni')
                        if selected_convocazione:
                            # Crea il percorso completo del file selezionato
                            file_path = os.path.join(acronimo, selected_convocazione)
                                
                            with open(file_path, 'r') as file:
                                convocazione = json.load(file)

                            # Form per inserire i dettagli della partita
                            giornata = st.number_input("Numero della Giornata", min_value=1, step=1)
                            squadra = st.text_input("Squadra Avversaria", convocazione['squadra'])
                            home_away = st.selectbox("Casa/Fuori Casa", ["Casa", "Fuori Casa"])
                            risultato = st.text_input("Risultato")
                            recupero = st.number_input("Eventuali minuti di recupero concesso", min_value=0, step=1)

                            # Formazione
                            st.subheader("Formazione")
                            formazione = []
                            available_players = convocazione["componenti_squadra"].copy()
                            for i in range(20):  # Supponiamo una formazione di 11 giocatori
                                selected_player = st.selectbox(f"N. {i + 1}", available_players, key=f"formazione_{i}")
                                formazione.append(selected_player)
                                available_players.remove(selected_player)

                            # Numero di sostituzioni
                            st.subheader(":blue[_Sostituzioni_]", divider="blue")
                            num_subs = st.number_input("Numero di Sostituzioni", min_value=0, max_value=11, step=1, key="subs")
                            # Sostituzioni
                            substitutions = []
                            for i in range(num_subs):
                                sub_in = st.selectbox(f"Giocatore sostituito {i + 1}", convocazione["componenti_squadra"], key=f"sub_in_{i}")
                                sub_out = st.selectbox(f"Giocatore subentrante {i + 1}", convocazione["componenti_squadra"], key=f"sub_out_{i}")
                                time_sub = st.number_input(f"Minuto della sostituzione {i + 1}", min_value=0, max_value=90, step=1, key=f"time_sub_{i}")
                                substitutions.append({"sub_in": sub_in, "sub_out": sub_out, "time_sub": time_sub})

                            # Ammonizioni
                            st.subheader(":orange[_Ammonizioni_]", divider="orange")
                            num_am = st.number_input("Numero di ammonizioni", min_value=0, max_value=20, step=1, key="amm")
                            ammonizioni = []
                            for i in range(num_am):  # Supponiamo un massimo di 5 ammonizioni
                                am_player = st.selectbox(f"Ammonizione {i + 1}", convocazione["componenti_squadra"], key=f"ammonizione_{i}")
                                ammonizioni.append(am_player)

                            # Espulsioni
                            st.subheader(":orange[_Espulsioni_]", divider="red")
                            num_esp = st.number_input("Numero di espulsioni", min_value=0, max_value=20, step=1, key="esp")
                            espulsioni = []
                            for i in range(num_esp):  # Supponiamo un massimo di 2 espulsioni
                                esp_player = st.selectbox(f"Espulsione {i + 1}", convocazione["componenti_squadra"], key=f"espulsione_{i}")
                                time_esp = st.number_input(f"Minuto dell'espulsione {i + 1}", min_value=0, max_value=100, step=1, key=f"time_esp_{i}")
                                espulsioni.append({"esp_player": esp_player, "time_esp": time_esp})

                            # Goal
                            st.subheader(":green[_Gol_]", divider="green")
                            num_gol = st.number_input("Numero di goal", min_value=0, max_value=20, step=1, key="gol")                   
                            goal = []
                            for i in range(num_gol):  # Supponiamo un massimo di 5 goal
                                goal_player = st.selectbox(f"Goal {i + 1}", convocazione["componenti_squadra"], key=f"goal_{i}")
                                goal.append(goal_player)

                            # Non Convocati
                            st.subheader(":gray[_Non Convocati_]", divider="gray")
                            non_convocati = []
                            motivi = ["INFORTUNATO", "SCELTA TECNICA", "MALATTIA", "NON ALLENATO", "NON DISPONIBILE", "RITIRATO", "ALTRA CATEGORIA"]
                            for player in convocazione["non_convocati"].split(", "):
                                motivo = st.selectbox(f"Motivo per {player}", motivi, key=f"motivo_{player}")
                                non_convocati.append({"giocatore": player, "motivo": motivo})

                            # Compila il dizionario
                            report_partita = {
                                "giornata": giornata,
                                "squadra": squadra,
                                "home_away": home_away,
                                "risultato": risultato,
                                "recupero": recupero,
                                "formazione": formazione,
                                "substitutions": substitutions,
                                "ammonizioni": ammonizioni,
                                "espulsioni": espulsioni,
                                "goal": goal,
                                "non_convocati": non_convocati
                            }

                            # Memorizza nel file JSON
                            if st.button("Salva Report"):
                                file_complete = os.path.join(selected_convocazione)
                                nfile, _ = os.path.splitext(file_complete)
                                report_match = reportistica + nfile + '.json'

                                st.write(report_match)
                                with open(report_match, 'w') as f:
                                    json.dump(report_partita, f, indent=4)
                                st.write("Report Partita:", report_partita)
                                st.success("Report salvato!")
                    else:
                        st.subheader("Non ci sono Convocazioni")

                elif selected=="Reportistica":
                    st.title("Report Statistiche")



                    def create_bar_chart(df, column, n_file, minuti):
                        # Crea un DataFrame per il grafico a barre
                        if column == 'presenze':
                            st.warning("⬅️ Visualizza il rapporto tra le giornate trascorse e le partite effettivamente giocate.")
                            #st.text("Visualizza il rapporto tra le giornate trascorse e le partite effettivamente giocate")
                            df['rapporto'] = (df[column])*100/(df['partite'])
                            chart_data = pd.DataFrame({
                                "Presenze": df['presenze'],
                                "Numero giornate": df['partite']
                            }, index=df['giocatore'])
                            chart_data = chart_data.sort_values(by="Presenze", ascending=False)
                            # Visualizza il grafico a barre
                            st.bar_chart(chart_data, horizontal=True)
                        elif column == 'minuti giocati':
                            st.warning("⬅️ Visualizza il rapporto tra i minuti totali giocabili e i minuti effettivamente giocati in %.")
                            minuti_totali = df['partite']*minuti
                            df['rapporto'] = (df['minuti giocati'])*100/(minuti_totali)
                            chart_data = pd.DataFrame({
                                'Minuti totali': minuti_totali,
                                'Rapporto (%) su minuti totali': df['rapporto']
                            }, index=df['giocatore'])
                            # Visualizza il grafico a barre
                            st.bar_chart(chart_data, horizontal=True)
                        elif column == 'generale':
                            st.warning("⬅️ Visualizza le partite da titolare, quelle in cui è stato sostituito (sub_out), quelle in cui è subentrato (sub_in), sulle presenze effettive.")
                            chart_data = pd.DataFrame(
                                {
                                    "giocatore": df['giocatore'],
                                    "partite": df['partite'],
                                    "presenze": df['presenze'],
                                    "titolare": df['titolare'],
                                    "sub_out": df['sub_out'],
                                    "sub_in": df['sub_in'],
                                }
                            )
                            st.bar_chart(chart_data, x="giocatore", y=["presenze", "titolare", "sub_out", "sub_in"], x_label="Informazioni presenze", color="partite", horizontal=True)
                        elif column == 'minuti giocati/convocazioni':
                            st.warning("⬅️ Visualizza il rapporto tra i minuti giocabili, basato sul numero di partite in cui è stato convocato, e i minuti effettivi giocati.")
                            convocazioni = df['partite'] - df['non convocazione']
                            minutaggio = minuti*convocazioni
                            df['rapporto'] = (df['minuti giocati'])*100/minutaggio
                            chart_data = pd.DataFrame(
                                {
                                    "giocatore": df['giocatore'],
                                    "convocazioni": convocazioni,
                                    "minutaggio": minutaggio,
                                    "Rapporto": df['rapporto']
                                }
                            )
                            st.bar_chart(chart_data, x="giocatore", y=["convocazioni", "minutaggio", "Rapporto"], x_label="Rapporto minuti giocati/totali convocazione", color="convocazioni")
                            
                        




                    # Funzione per creare un grafico a barre per il confronto tra due giocatori
                    def create_comparison_chart(df, player1, player2, column, title):
                        comparison_df = df[df['giocatore'].isin([player1, player2])]
                        st.markdown("""
                        <style>
                        .stText {
                            font-size: 20px;
                        }
                        </style>
                        """, unsafe_allow_html=True)
                        st.text(title)  
                        st.bar_chart(comparison_df.set_index('giocatore')[column])


                    def copy_df_to_xlsx(df_report, save_file, xlsx_file_path='report.xlsx'):
                        # Carica il file XLSX
                        workbook = openpyxl.load_workbook(xlsx_file_path)
                        sheet = workbook.active
                        
                        # Mappa delle colonne dal DataFrame all'XLSX
                        column_mapping = {
                            'giocatore': 'A',
                            'partite': 'B',
                            'presenze': 'C',
                            'titolare': 'D',
                            'sub_out': 'E',
                            'sub_in': 'F',
                            'minuti giocati': 'G',
                            'ammonizioni': 'I',
                            'espulsioni': 'J',
                            'goal': 'K',
                            'non convocazione': 'L'
                        }
                       
                        
                        # Inizia a scrivere i dati dalla riga 7
                        start_row = 7
                        for index, row in df_report.iterrows():
                            for df_col, xlsx_col in column_mapping.items():
                                cell_position = f"{xlsx_col}{str(start_row)}"
                                sheet[cell_position] = row[df_col]
                            start_row += 1
                        
                        # Salva il file XLSX
                        workbook.save(save_file)


                    giocatori = [f"{row['NOME']} {row['COGNOME']}" for index, row in df.iterrows()]

                    # Supponiamo che match_data_list sia una lista di dizionari JSON caricati dai file nella directory
                    match_data_list = []
                    n_file = 0
                    for filename in os.listdir(reportistica):
                        if filename.endswith('.json'):
                            with open(os.path.join(reportistica, filename), 'r') as file:
                                n_file += 1
                                match_data_list.append(json.load(file))

                    data = {}
                    # Itera attraverso tutti i file JSON nella lista
                    for match_data in match_data_list:
                        # Creazione delle liste richieste
                        first_11 = match_data["formazione"][:11]
                        sub_out = [sub["sub_out"] for sub in match_data["substitutions"]]
                        sub_in = [sub["sub_in"] for sub in match_data["substitutions"]]
                        time_sub = [sub["time_sub"] for sub in match_data["substitutions"]]
                        next_9 = match_data["formazione"][11:20]
                        non_convocati = [player["giocatore"] for player in match_data["non_convocati"]]
                        recupero = match_data['recupero']
                        #data['partite'] = n_file
                        # Aggiungi i giocatori della formazione
                        for player in first_11:
                            if player in sub_in:
                                if player not in data:
                                    time_play = time_sub[sub_in.index(player)] + recupero
                                    data[player] = {
                                        "presenze": 1,
                                        "giocatore": player,
                                        "titolare": 1,
                                        "sub_out": 1,
                                        "sub_in": 0,
                                        "minuti giocati": time_play,
                                        "ammonizioni": 0,
                                        "espulsioni": 0,
                                        "goal": 0,
                                        "non convocazione": 0
                                    }
                                else:
                                    time_play = time_sub[sub_in.index(player)] + recupero
                                    data[player]["minuti giocati"] += time_play
                                    data[player]["presenze"] += 1
                                    data[player]["titolare"] += 1
                                    data[player]["sub_out"] += 1
                            else:
                                if player not in data:
                                    time_play = minuti + recupero
                                    data[player] = {
                                        "presenze": 1,
                                        "giocatore": player,
                                        "titolare": 1,
                                        "sub_out": 0,
                                        "sub_in": 0,                                
                                        "minuti giocati": time_play,
                                        "ammonizioni": 0,
                                        "espulsioni": 0,
                                        "goal": 0,
                                        "non convocazione": 0
                                    }
                                else:
                                    time_play = minuti + recupero
                                    data[player]["minuti giocati"] += time_play
                                    data[player]["presenze"] += 1
                                    data[player]["titolare"] += 1

                        # Aggiungi i giocatori dalla panchina
                        for player in next_9:
                            if player in sub_out and player in sub_in:
                                time_play = time_sub[sub_in.index(player)] - time_sub[sub_out.index(player)]
                                if player not in data:
                                    data[player] = {
                                        "presenze": 1,
                                        "giocatore": player,
                                        "titolare": 0,
                                        "sub_out": 1,
                                        "sub_in": 1,
                                        "minuti giocati": time_play,
                                        "ammonizioni": 0,
                                        "espulsioni": 0,
                                        "goal": 0,
                                        "non convocazione": 0
                                    }
                                else:
                                    data[player]["minuti giocati"] += time_play
                                    data[player]["presenze"] += 1
                                    data[player]["sub_out"] += 1
                                    data[player]["sub_in"] += 1
                            elif player in sub_out:
                                if player not in data:
                                    time_play = minuti - time_sub[sub_out.index(player)] + recupero
                                    data[player] = {
                                        "presenze": 1,
                                        "giocatore": player,
                                        "titolare": 0,
                                        "sub_out": 0,
                                        "sub_in": 1,
                                        "minuti giocati": time_play,
                                        "ammonizioni": 0,
                                        "espulsioni": 0,
                                        "goal": 0,
                                        "non convocazione": 0
                                    }
                                else:
                                    time_play = minuti - time_sub[sub_out.index(player)] + recupero
                                    data[player]["minuti giocati"] += time_play
                                    data[player]["presenze"] += 1
                                    data[player]["sub_in"] += 1
                            else:
                                if player not in data:
                                    data[player] = {
                                        "presenze": 0,
                                        "giocatore": player,
                                        "titolare": 0,
                                        "sub_out": 0,
                                        "sub_in": 0,                                
                                        "minuti giocati": 0,
                                        "ammonizioni": 0,
                                        "espulsioni": 0,
                                        "goal": 0,
                                        "non convocazione": 0
                                    }
                                else:
                                    data[player]["minuti giocati"] += 0
                                    data[player]["presenze"] += 0


                        # Aggiungi i giocatori ammoniti
                        for player in match_data["ammonizioni"]:
                            if player not in data:
                                data[player] = {"presenze": 0, "giocatore": player, "titolare": 0, "sub_out": 0, "sub_in": 0, "minuti giocati": 0, "ammonizioni": 1, "espulsioni": 0, "goal": 0, "non convocazione": 0}
                            else:
                                data[player]["ammonizioni"] += 1

                        # Aggiungi i giocatori espulsi
                        for player in match_data["espulsioni"]:
                            esp_player = player['esp_player']
                            time_esp = player['time_esp']
                                            
                            if esp_player in sub_in:
                                if esp_player not in data:
                                    min_giocati = time_sub[sub_in.index(player)] - time_esp
                                    data[esp_player] = {
                                        "presenze": 1,
                                        "giocatore": esp_player,
                                        "titolare": 1,
                                        "sub_out": 1,
                                        "sub_in": 0,
                                        "minuti giocati": min_giocati,
                                        "ammonizioni": 0,
                                        "espulsioni": 1,
                                        "goal": 0,
                                        "non convocazione": 0
                                    }
                                else:
                                    min_giocati = data[esp_player]["minuti giocati"] - time_sub[sub_in.index(player)]
                                    data[esp_player]["minuti giocati"] += (min_giocati)
                                    data[esp_player]["espulsioni"] += 1
                            else:

                                data[esp_player]["espulsioni"] += 1
                                min_giocati = data[esp_player]["minuti giocati"] - minuti
                                #st.text(min_giocati)
                                data[esp_player]["minuti giocati"] = (min_giocati + time_esp)

                        # Aggiungi i giocatori che hanno segnato
                        for player in match_data["goal"]:
                            if player not in data:
                                data[player] = {"presenze": 0, "giocatore": player, "minuti giocati": 0, "titolare": 0, "sub_out": 0, "sub_in": 0, "ammonizioni": 0, "espulsioni": 0, "goal": 1, "non convocazione": 0}
                            else:
                                data[player]["goal"] += 1

                        # Aggiungi i giocatori non convocati
                        for non_convocato in match_data["non_convocati"]:
                            player = non_convocato["giocatore"]
                            if player not in data:
                                data[player] = {"presenze": 0, "giocatore": player, "minuti giocati": 0, "titolare": 0, "sub_out": 0, "sub_in": 0, "ammonizioni": 0, "espulsioni": 0, "goal": 0, "non convocazione": 1}
                            else:
                                data[player]["non convocazione"] += 1


                    # Creazione del dataframe
                    df_report = pd.DataFrame.from_dict(data, orient='index', columns=["giocatore", "presenze", "titolare", "sub_out", "sub_in", "minuti giocati", "ammonizioni", "espulsioni", "goal", "non convocazione"])
                    df_report.insert(1, 'partite', n_file)

                    # Visualizza la tabella
                    st.subheader("Tabella Giocatori")
                    st.dataframe(df_report.reset_index(drop=True), use_container_width=True)

                    # Salvataggio del dataframe in formato csv
                    # file_path = reportistica + 'reportistica_' + mister_info['acronimo'] + '.csv'

                    #Esporta la reportistica in excel
                    fname = reportistica + 'reportistica.xlsx'
                    copy_df_to_xlsx(df_report, fname)


                    # Salva il file Excel nella cartella convocazioni
                    
                    # convert_df_to_excel(df_report, fname)

                    # Genera il link di download per il file Excel
                    with open(fname, 'rb') as f:
                        excel_buffer = BytesIO(f.read())
                        download_link_html = download_link(excel_buffer.getvalue(), fname, 'Clicca qui per scaricare il file Excel')
                        commit_message = "Update file"
                        st.success("Scarica tabella in Excel")

                        st.markdown(download_link_html, unsafe_allow_html=True)


                    # Creazione dei grafici
                    # Seleziona la statistica da visualizzare
                    
                    # Calcola il rapporto
                    stat = st.selectbox('Seleziona la statistica da visualizzare', 
                                            ['presenze', 'minuti giocati', 'minuti giocati/convocazioni', 'generale'])
                    df_report['minuti_p'] = ((df_report['minuti giocati'])*100/(minuti * n_file)).apply(lambda x: f"{x:,.2f}".replace('.', ','))
                    df_report['presenze_p'] = ((df_report['presenze'])*100/(df_report['presenze']) * 100).apply(lambda x: f"{x:,.2f}".replace('.', ','))
                    # Crea il grafico a barre per la statistica selezionata
                    create_bar_chart(df_report, stat, n_file, minuti)
                    # df_report.to_csv(file_path, sep=";",  index=False)

                    # Seleziona i giocatori da confrontare
                    statistica = st.selectbox('Seleziona la statistica da visualizzare', 
                                            ['presenze', 'minuti giocati', 'minuti giocati/convocazioni', 'generale'], key='stat')            
                    # player1 = st.selectbox('Seleziona il primo giocatore', df_report['giocatore'])
                    # player2 = st.selectbox('Seleziona il secondo giocatore', df_report['giocatore'])

                    # # # Crea il grafico a barre per la statistica selezionata
                    # if player1 and player2:
                    #     create_comparison_chart(df_report, player1, player2, statistica, f'Confronto {statistica} tra {player1} e {player2}')
                    # else:
                    #     create_bar_chart(df_report, statistica, f'{statistica.capitalize()} dei Giocatori')


                    # Seleziona fino a tre giocatori per il confronto
                    giocatori = df_report['giocatore'].unique()
                    selezionati = st.multiselect('Seleziona fino a tre giocatori', giocatori, placeholder="scegli...", max_selections=4)

                    # Filtra i dati per i giocatori selezionati
                    if len(selezionati) > 0:
                        df_selected = df_report[df_report['giocatore'].isin(selezionati)]
                        create_bar_chart(df_selected, statistica, n_file, minuti)
                    else:
                        st.write("Seleziona almeno un giocatore per visualizzare il grafico.")

def impostazioni():
    # Esempio di impostazioni
    tema = st.selectbox("Seleziona il tema", ["Chiaro", "Scuro"])
    notifiche = st.checkbox("Abilita notifiche")
    st.write(f"Tema selezionato: {tema}")
    st.write(f"Notifiche abilitate: {notifiche}")


# ---------- SIDEBAR ----------
# Navigazione tra le pagine
pages = {
    "Scuola Calcio": homepage,
    "Prima Squadra e SGS": gestione_rosa,
    "Impostazioni": impostazioni
}

with st.sidebar:
    selection = option_menu("Main Menu", ["Scuola Calcio", "Prima Squadra e SGS",'Impostazioni'], 
        icons=['', '', 'gear'], menu_icon="cast", default_index=1)


    social_media_links = [
        "https://www.facebook.com/ThisIsAnExampleLink",
        "https://www.instagram.com/ThisIsAnExampleLink",
    ]
    colors = ["#000000", None, "SteelBlue", None]
    social_media_icons = SocialMediaIcons(social_media_links, colors)
    social_media_icons.render()
# Caricamento della pagina selezionata
page = pages[selection]
page()